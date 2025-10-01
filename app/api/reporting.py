import io
import json
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from flask_sqlalchemy import SQLAlchemy

class ReportGenerator:
    def __init__(self, db: SQLAlchemy):
        self.db = db
        
    def generate_team_performance_report(self, team_id: int, start_date: Optional[str] = None, 
                                       end_date: Optional[str] = None, format: str = 'pdf') -> io.BytesIO:
        """Generate comprehensive team performance report"""
        
        # Import models here to avoid circular imports
        from app import Team, Game, PlayData, Visualization
        
        # Get team data
        team = Team.query.get(team_id)
        if not team:
            raise ValueError("Team not found")
        
        # Filter games by date range if provided
        query = Game.query.filter_by(team_id=team_id)
        if start_date:
            query = query.filter(Game.submission_timestamp >= start_date)
        if end_date:
            query = query.filter(Game.submission_timestamp <= end_date)
        
        games = query.order_by(Game.week).all()
        
        if format == 'pdf':
            return self._generate_pdf_report(team, games)
        elif format == 'excel':
            return self._generate_excel_report(team, games)
        else:
            raise ValueError("Unsupported format")
    
    def _generate_pdf_report(self, team, games) -> io.BytesIO:
        """Generate PDF performance report"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1*inch)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.darkblue,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.darkblue,
            spaceBefore=20,
            spaceAfter=10
        )
        
        # Story elements
        story = []
        
        # Title
        story.append(Paragraph(f"{team.team_name} Performance Report", title_style))
        story.append(Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", heading_style))
        
        if games:
            total_plays = sum(len(game.play_data) for game in games)
            total_yards = sum(sum(play.yards_gained for play in game.play_data) for game in games)
            avg_yards_per_play = total_yards / total_plays if total_plays > 0 else 0
            
            summary_data = [
                ['Metric', 'Value'],
                ['Total Games Analyzed', str(len(games))],
                ['Total Plays', str(total_plays)],
                ['Total Yards Gained', str(total_yards)],
                ['Average Yards per Play', f"{avg_yards_per_play:.2f}"],
                ['Date Range', f"Week {min(g.week for g in games)} - Week {max(g.week for g in games)}"]
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 20))
            
            # Game-by-Game Analysis
            story.append(Paragraph("Game-by-Game Analysis", heading_style))
            
            for game in games:
                story.append(Paragraph(f"Week {game.week} vs {game.opponent} ({game.location})", styles['Heading3']))
                
                plays = game.play_data
                game_yards = sum(play.yards_gained for play in plays)
                game_points = sum(play.points_scored for play in plays)
                
                # Formation breakdown
                formations = {}
                for play in plays:
                    formations[play.formation] = formations.get(play.formation, 0) + 1
                
                top_formation = max(formations.items(), key=lambda x: x[1]) if formations else ("N/A", 0)
                
                game_data = [
                    ['Total Plays', str(len(plays))],
                    ['Total Yards', str(game_yards)],
                    ['Total Points', str(game_points)],
                    ['Avg Yards/Play', f"{game_yards/len(plays):.2f}" if plays else "0"],
                    ['Most Used Formation', f"{top_formation[0]} ({top_formation[1]} plays)"],
                ]
                
                if game.analytics_focus_notes:
                    story.append(Paragraph(f"Focus: {game.analytics_focus_notes}", styles['Italic']))
                
                game_table = Table(game_data, colWidths=[2*inch, 2*inch])
                game_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(game_table)
                story.append(Spacer(1, 15))
            
            # Trends and Insights
            story.append(PageBreak())
            story.append(Paragraph("Performance Trends", heading_style))
            
            # Weekly performance data
            weekly_data = [['Week', 'Opponent', 'Total Yards', 'Points', 'Avg Yards/Play']]
            for game in games:
                plays = game.play_data
                yards = sum(play.yards_gained for play in plays)
                points = sum(play.points_scored for play in plays)
                avg = yards / len(plays) if plays else 0
                weekly_data.append([
                    str(game.week),
                    game.opponent,
                    str(yards),
                    str(points),
                    f"{avg:.2f}"
                ])
            
            weekly_table = Table(weekly_data, colWidths=[0.8*inch, 1.5*inch, 1*inch, 1*inch, 1.2*inch])
            weekly_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(weekly_table)
            
        else:
            story.append(Paragraph("No games found for the specified criteria.", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def _generate_excel_report(self, team, games) -> io.BytesIO:
        """Generate Excel performance report with charts"""
        buffer = io.BytesIO()
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Summary sheet
        summary_sheet = workbook.create_sheet("Summary", 0)
        self._create_summary_sheet(summary_sheet, team, games)
        
        # Game details sheet
        details_sheet = workbook.create_sheet("Game Details", 1)
        self._create_game_details_sheet(details_sheet, games)
        
        # Charts sheet
        charts_sheet = workbook.create_sheet("Charts", 2)
        self._create_charts_sheet(charts_sheet, games)
        
        # Raw data sheet
        raw_sheet = workbook.create_sheet("Raw Data", 3)
        self._create_raw_data_sheet(raw_sheet, games)
        
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _create_summary_sheet(self, sheet, team, games):
        """Create Excel summary sheet"""
        # Title
        sheet['A1'] = f"{team.team_name} Performance Summary"
        sheet['A1'].font = Font(size=18, bold=True, color="1F4E79")
        sheet.merge_cells('A1:D1')
        
        # Metadata
        sheet['A3'] = "Report Generated:"
        sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sheet['A4'] = "Total Games:"
        sheet['B4'] = len(games)
        
        if games:
            # Calculate metrics
            total_plays = sum(len(game.play_data) for game in games)
            total_yards = sum(sum(play.yards_gained for play in game.play_data) for game in games)
            total_points = sum(sum(play.points_scored for play in game.play_data) for game in games)
            
            # Metrics table
            headers = ['Metric', 'Value', 'Average per Game']
            metrics = [
                ['Total Plays', total_plays, total_plays / len(games) if games else 0],
                ['Total Yards', total_yards, total_yards / len(games) if games else 0],
                ['Total Points', total_points, total_points / len(games) if games else 0],
                ['Yards per Play', total_yards / total_plays if total_plays > 0 else 0, ''],
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=6, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Write metrics
            for row, metric in enumerate(metrics, 7):
                for col, value in enumerate(metric, 1):
                    if col == 3 and isinstance(value, (int, float)) and value != '':
                        sheet.cell(row=row, column=col, value=round(value, 2))
                    else:
                        sheet.cell(row=row, column=col, value=value)
            
            # Auto-adjust column widths
            for col in range(1, 4):
                sheet.column_dimensions[chr(64 + col)].width = 20
    
    def _create_game_details_sheet(self, sheet, games):
        """Create game-by-game details sheet"""
        headers = ['Week', 'Opponent', 'Location', 'Total Plays', 'Total Yards', 'Total Points', 'Avg Yards/Play', 'Top Formation']
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Write game data
        for row, game in enumerate(games, 2):
            plays = game.play_data
            yards = sum(play.yards_gained for play in plays)
            points = sum(play.points_scored for play in plays)
            
            # Get top formation
            formations = {}
            for play in plays:
                formations[play.formation] = formations.get(play.formation, 0) + 1
            top_formation = max(formations.items(), key=lambda x: x[1])[0] if formations else "N/A"
            
            data = [
                game.week,
                game.opponent,
                game.location,
                len(plays),
                yards,
                points,
                round(yards / len(plays), 2) if plays else 0,
                top_formation
            ]
            
            for col, value in enumerate(data, 1):
                sheet.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[chr(64 + col)].width = 15
    
    def _create_charts_sheet(self, sheet, games):
        """Create charts sheet with performance visualizations"""
        if not games:
            sheet['A1'] = "No data available for charts"
            return
        
        # Weekly yards chart
        sheet['A1'] = "Weekly Performance Trends"
        sheet['A1'].font = Font(size=14, bold=True)
        
        # Prepare data for chart
        chart_data = [['Week', 'Total Yards', 'Total Points']]
        for game in games:
            plays = game.play_data
            yards = sum(play.yards_gained for play in plays)
            points = sum(play.points_scored for play in plays)
            chart_data.append([f"Week {game.week}", yards, points])
        
        # Write chart data
        for row, data_row in enumerate(chart_data, 3):
            for col, value in enumerate(data_row, 1):
                sheet.cell(row=row, column=col, value=value)
        
        # Create line chart
        chart = LineChart()
        chart.title = "Weekly Performance"
        chart.y_axis.title = 'Yards/Points'
        chart.x_axis.title = 'Week'
        
        data = Reference(sheet, min_col=2, min_row=3, max_col=3, max_row=3 + len(games))
        cats = Reference(sheet, min_col=1, min_row=4, max_row=3 + len(games))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        sheet.add_chart(chart, "E3")
    
    def _create_raw_data_sheet(self, sheet, games):
        """Create raw play-by-play data sheet"""
        headers = ['Game Week', 'Opponent', 'Play ID', 'Down', 'Distance', 'Yard Line', 
                  'Formation', 'Play Type', 'Play Name', 'Result', 'Yards Gained', 'Points Scored']
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Write play data
        row = 2
        for game in games:
            for play in game.play_data:
                data = [
                    game.week,
                    game.opponent,
                    play.play_id,
                    play.down,
                    play.distance,
                    play.yard_line,
                    play.formation,
                    play.play_type,
                    play.play_name,
                    play.result_of_play,
                    play.yards_gained,
                    play.points_scored
                ]
                
                for col, value in enumerate(data, 1):
                    sheet.cell(row=row, column=col, value=value)
                row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[chr(64 + col)].width = 15
    
    def generate_consultant_report(self, consultant_id: int, team_ids: List[int], 
                                 format: str = 'pdf') -> io.BytesIO:
        """Generate consultant performance report across multiple teams"""
        
        from app import Consultant, Team, Game, PlayData, Visualization
        
        consultant = Consultant.query.get(consultant_id)
        if not consultant:
            raise ValueError("Consultant not found")
        
        teams = Team.query.filter(Team.id.in_(team_ids)).all()
        
        if format == 'pdf':
            return self._generate_consultant_pdf_report(consultant, teams)
        elif format == 'excel':
            return self._generate_consultant_excel_report(consultant, teams)
        else:
            raise ValueError("Unsupported format")
    
    def _generate_consultant_pdf_report(self, consultant, teams) -> io.BytesIO:
        """Generate PDF consultant report"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1*inch)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.darkgreen,
            spaceAfter=30,
            alignment=1
        )
        
        story = []
        
        # Title
        story.append(Paragraph(f"Consultant Report - {consultant.name}", title_style))
        story.append(Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Teams overview
        team_data = [['Team Name', 'Total Games', 'Total Plays', 'Avg Yards/Game']]
        
        for team in teams:
            games = Game.query.filter_by(team_id=team.id).all()
            total_plays = sum(len(game.play_data) for game in games)
            total_yards = sum(sum(play.yards_gained for play in game.play_data) for game in games)
            avg_yards = total_yards / len(games) if games else 0
            
            team_data.append([
                team.team_name,
                str(len(games)),
                str(total_plays),
                f"{avg_yards:.1f}"
            ])
        
        team_table = Table(team_data, colWidths=[2*inch, 1*inch, 1*inch, 1.5*inch])
        team_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(team_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def _generate_consultant_excel_report(self, consultant, teams) -> io.BytesIO:
        """Generate Excel consultant report"""
        buffer = io.BytesIO()
        workbook = openpyxl.Workbook()
        
        # Implementation similar to team report but for multiple teams
        # ... (abbreviated for brevity)
        
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

# Initialize global instance
report_generator = None

def init_reporting(db):
    global report_generator
    report_generator = ReportGenerator(db)